import os
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, numbers

def count_words_in_file(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return len(file.read().split())
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return 0

def calculate_reward_multiplier(lines, start_idx):
    """
    Calculate reward multiplier based on number of contributors
    between contributor_names: and reward:
    """
    contributor_count = 0
    idx = start_idx + 1  # Start from next line after contributors_id
    
    while idx < len(lines) and not lines[idx].strip().startswith('reward:'):
        if lines[idx].strip().startswith('-'):
            contributor_count += 1
        idx += 1
    
    if contributor_count == 1:
        return 2
    elif contributor_count >= 2:  # Changed to handle multiple contributors
        return 4
    return 1

def read_yaml_file(yaml_path):
    """
    Read the yaml file (course.yml or tutorial.yml) and extract all language-reward pairs,
    applying multipliers based on contributor count.
    """
    try:
        with open(yaml_path, 'r', encoding='utf-8') as file:
            lines = [line.rstrip() for line in file.readlines()]
            language_rewards = {}
            current_language = None
            i = 0
            
            while i < len(lines):
                line = lines[i].strip()
                
                if line.startswith('- language:'):
                    current_language = line.split('- language:')[1].strip()
                    i += 1  # Move to the next line
                    continue
                
                if line.startswith('contributor_names:') and current_language:
                    # Find reward multiplier based on contributors
                    multiplier = calculate_reward_multiplier(lines, i)
                    
                    # Find the reward value
                    reward_idx = i
                    while reward_idx < len(lines):
                        if lines[reward_idx].strip().startswith('reward:'):
                            try:
                                base_reward = float(lines[reward_idx].split('reward:')[1].strip())
                                final_reward = base_reward * multiplier
                                language_rewards[current_language] = final_reward
                                print(f"Language: {current_language}, Base Reward: ${base_reward:.2f}, "
                                      f"Multiplier: {multiplier}, Final Reward: ${final_reward:.2f}")
                                break
                            except ValueError:
                                print(f"Warning: Invalid reward value in {yaml_path} for language {current_language}")
                                break
                        reward_idx += 1
                    
                    # Move to the next language section
                    current_language = None
                
                # If we encounter a reward directly (without contributors_id)
                elif line.startswith('reward:') and current_language:
                    try:
                        base_reward = float(line.split('reward:')[1].strip())
                        # No multiplier in this case
                        language_rewards[current_language] = base_reward
                        print(f"Language: {current_language}, Reward: ${base_reward:.2f} (no multiplier)")
                        
                        # Reset current_language
                        current_language = None
                    except ValueError:
                        print(f"Warning: Invalid reward value in {yaml_path} for language {current_language}")
                
                i += 1
                        
            return language_rewards
    except Exception as e:
        print(f"Error reading YAML {yaml_path}: {e}")
        return {}

def process_courses(courses_path):
    results = []
    languages = set()  # Set to store unique languages
    course_data = []  # List to store course data
    
    # Check if base_path exists
    if not courses_path.exists():
        print(f"Warning: Courses path {courses_path} does not exist")
        return [], set()
    
    print(f"Looking for courses in: {courses_path}")
    
    # First pass: collect all languages and course data
    for folder in courses_path.iterdir():
        if folder.is_dir():
            yaml_file = folder / 'course.yml'
            md_file = folder / 'en.md'
            
            if yaml_file.exists():
                print(f"\nProcessing {yaml_file}")
                language_rewards = read_yaml_file(yaml_file)
                languages.update(language_rewards.keys())
                
                course_data.append({
                    'Folder': folder.name,
                    'Word Count': count_words_in_file(md_file) if md_file.exists() else 0,
                    'language_rewards': language_rewards
                })
    
    # Print all found languages for verification
    print(f"\nFound {len(course_data)} courses")
    print("All found languages:", sorted(languages))
    
    # Create the final data structure
    for course in course_data:
        row_data = {
            'Folder': course['Folder'],
            'Word Count': course['Word Count']
        }
        # Add the reward under the corresponding language column
        for lang in languages:
            row_data[lang] = course['language_rewards'].get(lang)
        
        results.append(row_data)
    
    return results, languages

def process_tutorials(tutorials_path):
    results = []
    languages = set()  # Set to store unique languages
    tutorial_data = []  # List to store tutorial data
    
    # Check if base_path exists
    if not tutorials_path.exists():
        print(f"Warning: Tutorials path {tutorials_path} does not exist")
        return [], set()
    
    print(f"Looking for tutorials in: {tutorials_path}")
    tutorial_count = 0
    
    # Process level 2 directories (categories)
    for category_dir in tutorials_path.iterdir():
        if category_dir.is_dir():
            category_name = category_dir.name
            print(f"\nScanning category: {category_name}")
            
            # Process level 3 directories (tutorial names)
            for tutorial_dir in category_dir.iterdir():
                if tutorial_dir.is_dir():
                    yaml_file = tutorial_dir / 'tutorial.yml'
                    md_file = tutorial_dir / 'en.md'
                    
                    if yaml_file.exists():
                        tutorial_count += 1
                        print(f"  Processing tutorial: {tutorial_dir.name}")
                        language_rewards = read_yaml_file(yaml_file)
                        languages.update(language_rewards.keys())
                        
                        # Use both category and tutorial name for better identification
                        folder_name = f"{category_name}/{tutorial_dir.name}"
                        
                        tutorial_data.append({
                            'Folder': folder_name,
                            'Word Count': count_words_in_file(md_file) if md_file.exists() else 0,
                            'language_rewards': language_rewards
                        })
    
    print(f"\nFound {tutorial_count} tutorials")
    print("All found languages in tutorials:", sorted(languages))
    
    # Create the final data structure
    for tutorial in tutorial_data:
        row_data = {
            'Folder': tutorial['Folder'],
            'Word Count': tutorial['Word Count']
        }
        # Add the reward under the corresponding language column
        for lang in languages:
            row_data[lang] = tutorial['language_rewards'].get(lang)
        
        results.append(row_data)
    
    return results, languages

def main():
    try:
        # Use the direct relative paths as specified
        script_dir = Path(__file__).resolve().parent
        courses_path = script_dir / '../../../courses'
        tutorials_path = script_dir / '../../../tutorials'
        
        print(f"Script directory: {script_dir.resolve()}")
        print(f"Looking for courses in: {courses_path.resolve()}")
        print(f"Looking for tutorials in: {tutorials_path.resolve()}")
        
        # Process courses and tutorials
        courses_results, course_languages = process_courses(courses_path)
        tutorials_results, tutorial_languages = process_tutorials(tutorials_path)
        
        # Combine results
        results = []
        results.extend(courses_results)
        results.extend(tutorials_results)
        
        # Combine all found languages
        all_languages = course_languages.union(tutorial_languages)
        
        if not results:
            print("No data found. Please check if the paths are correct.")
            return
        
        # Save results in the same directory as the script
        output_path = script_dir / 'word_counts.xlsx'
        
        # Create DataFrame
        df = pd.DataFrame(results)
        
        if 'Word Count' in df.columns:
            # Rename 'Word Count' to 'Word Count EN'
            df = df.rename(columns={'Word Count': 'Word Count EN'})
        elif not df.empty and 'Word Count EN' not in df.columns:
            # If neither column exists, add Word Count EN
            df['Word Count EN'] = None
        
        # Ensure all language columns exist
        for lang in all_languages:
            if lang not in df.columns:
                df[lang] = None
        
        # Sort the columns to ensure Folder and Word Count EN are first, followed by alphabetically sorted languages
        columns = ['Folder', 'Word Count EN'] + sorted([col for col in df.columns if col not in ['Folder', 'Word Count EN']])
        df = df[columns]
        
        # Calculate totals for proofreading
        first_proofreading = {}
        first_proofreading['Folder'] = 'Total for first proofreading'
        first_proofreading['Word Count EN'] = None
        for col in columns[2:]:  # Skip Folder and Word Count EN
            first_proofreading[col] = df[col].sum()
        
        # Add empty row
        empty_row = {col: None for col in columns}
        
        # Calculate second proofreading
        second_proofreading = {}
        second_proofreading['Folder'] = 'Total for second proofreading'
        second_proofreading['Word Count EN'] = None
        for col in columns[2:]:
            second_proofreading[col] = first_proofreading[col] / 2

        # Calculate total for all proofreading (2 rounds now)
        total_all = {}
        total_all['Folder'] = 'Total for all the proofreading'
        total_all['Word Count EN'] = None
        for col in columns[2:]:
            total_all[col] = first_proofreading[col] + second_proofreading[col]

        # Append all new rows to the DataFrame
        df = pd.concat([df,
                       pd.DataFrame([first_proofreading]),
                       pd.DataFrame([empty_row]),
                       pd.DataFrame([second_proofreading]),
                       pd.DataFrame([empty_row]),
                       pd.DataFrame([total_all])],
                      ignore_index=True)
        
        # Create Excel file with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write the DataFrame to Excel
            df.to_excel(writer, index=False)
            
            # Get the worksheet
            worksheet = writer.sheets['Sheet1']
            
            # Apply dollar formatting to all language columns (excluding Folder and Word Count EN)
            for idx, column_name in enumerate(columns[2:], start=3):  # Start from 3rd column (index 2 + 1)
                col_letter = openpyxl.utils.get_column_letter(idx)
                for row in range(2, worksheet.max_row + 1):  # Start from second row (skip header)
                    cell = worksheet[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'
            
            # Find the row index of "Total for all the proofreading"
            total_row = None
            for idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row), 1):
                if row[0].value == 'Total for all the proofreading':
                    total_row = idx
                    break
            
            # Apply bold formatting to the entire row
            if total_row:
                for cell in worksheet[total_row]:
                    cell.font = Font(bold=True)

        print(f"\nResults saved to {output_path}")
        
    except FileNotFoundError as e:
        print("Error: Make sure this script is placed in the correct location relative to courses and tutorials folders")
        print(e)
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
